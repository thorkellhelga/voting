#coding:utf-8
import random
from backports import csv
from math import log, sqrt
import sys #??????
from tabulate import tabulate
import io
import xlsxwriter
import openpyxl
from copy import deepcopy, copy
import configparser
import codecs

from methods import var_alt_scal, alternating_scaling, icelandic_law
from methods import monge, nearest_neighbor, relative_superiority
from methods import norwegian_law, norwegian_icelandic
from methods import opt_entropy, switching
from methods import pure_vote_ratios

#??????
def random_id(length=8):
    chars = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
    s = "".join(random.sample(chars, length))
    return s


def read_csv(filename):
    with io.open(filename, mode="r", newline='', encoding='utf-8') as f:
        for row in csv.reader(f, skipinitialspace=True):
            yield [cell for cell in row]

def read_xlsx(filename):
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    for row in sheet.rows:
        yield [cell.value for cell in row]

def load_constituencies(confile):
    """Load constituencies from a file."""
    if confile.endswith("csv"):
        reader = read_csv(confile)
    else:
        reader = read_xlsx(confile)
    cons = []
    for row in reader:
        try:
            assert(int(row[1]) + int(row[2]) >= 0)
        except Exception as e:
            print(row[1:3])
            raise Exception("Error loading constituency file: "
                            "constituency seats and adjustment seats "
                            "must add to a nonzero number.")
        cons.append({
            "name": row[0],
            "num_constituency_seats": int(row[1]),
            "num_adjustment_seats": int(row[2])})
    return cons

def load_votes_from_stream(stream, filename):
    rd = []
    if filename.endswith(".csv"):
        if isinstance(stream, io.BytesIO):
            stream = codecs.iterdecode(stream, 'utf-8')
        for row in csv.reader(stream, skipinitialspace=True):
            rd.append(row)
    elif filename.endswith(".xlsx"):
        book = openpyxl.load_workbook(stream)
        sheet = book.active
        for row in sheet.rows:
            rd.append([cell.value for cell in row])
    else:
        return None, None, None

    const_seats_incl = rd[0][1].lower() == "cons"
    expected = 2 if const_seats_incl else 1
    adj_seats_incl = rd[0][expected].lower() == "adj"

    return parse_input(
        input=rd,
        parties_included=True,
        const_included=True,
        const_seats_included=const_seats_incl,
        adj_seats_included=adj_seats_incl
    )

def parse_input(
    input,
    parties_included,
    const_included,
    const_seats_included,
    adj_seats_included
):
    res = {}
    if parties_included:
        res["parties"] = input[0]
        del(input[0])

    if const_included:
        res["constituencies"] = [row[0] for row in input]
        for row in input: del(row[0])
        if parties_included: res["parties"] = res["parties"][1:]

    if const_seats_included:
        res["constituency_seats"] = [parsint(row[0]) for row in input]
        for row in input: del(row[0])
        if parties_included: res["parties"] = res["parties"][1:]

    if adj_seats_included:
        res["constituency_adjustment_seats"] = [parsint(row[0]) for row in input]
        for row in input: del(row[0])
        if parties_included: res["parties"] = res["parties"][1:]

    res["votes"] = input
    if parties_included:
        while not res["parties"][-1]:
            res["parties"] = res["parties"][:-1]
        res["votes"] = [row[:len(res["parties"])] for row in res["votes"]]
    res["votes"] = [[parsint(v) for v in row] for row in res["votes"]]
    return res

def parsint(value):
    return int(value) if value else 0


def load_votes(votefile, consts):
    """Load votes from a file."""
    if votefile.endswith("csv"):
        reader = read_csv(votefile)
    else:
        reader = read_xlsx(votefile)
    parties = next(reader)[3:]
    votes = [[] for i in range(len(consts))]
    c_names = [x["name"] for x in consts]

    for row in reader:
        try:
            v = votes[c_names.index(row[0])]
        except:
            print(row)
            raise Exception("Constituency '%s' not found in constituency file"
                            % row[0])
        for x in row[3:]:
            try:
                r = float(x)
            except:
                r = 0
            v.append(r)

    return parties, votes

def add_totals(m):
    """Add sums of rows and columns to a table."""
    nm = deepcopy(m)
    for i in range(len(m)):
        nm[i].append(sum(m[i]))
    totals = [sum(x) for x in zip(*nm)]
    nm.append(totals)
    return nm

def matrix_subtraction(A, B):
    m = len(A)
    assert(len(B) == m)
    if m == 0:
        return []
    n = len(A[0])
    assert(all([len(A[i]) == n and len(B[i]) == n for i in range(m)]))
    return [
        [A[i][j] - B[i][j] for j in range(n)]
        for i in range(m)
    ]

def find_xtd_shares(xtd_table):
    return [[float(v)/c[-1] if c[-1]!=0 else 0 for v in c] for c in xtd_table]

def find_shares(table):
    return [[float(v)/sum(c) if sum(c)!=0 else 0 for v in c] for c in table]

def print_table(data, header, labels, output, f_string=None):
    """
    Print 'data' in a table with 'header' and rows labelled with 'labels'.
    """
    if f_string:
        data = [[f_string.format(d) if d!=None else d for d in row] for row in data]
    data = [[labels[i]] + data[i] for i in range(len(data))]
    data = [[d if d != 0 and d != "0.0%" else None for d in row]
                for row in data]
    print(tabulate(data, header, output))

def print_steps_election(election):
    """Print detailed information about a single election."""
    rules = election.rules
    out = rules["output"]
    header = ["Constituency"]
    header.extend(rules["parties"])
    header.append("Total")
    if "constituencies" in rules:
        const_names = [c["name"] for c in rules["constituencies"]]
    else:
        const_names = rules["constituency_names"]
    const_names.append("Total")

    print("Votes")
    xtd_votes = add_totals(election.m_votes)
    print_table(xtd_votes, header, const_names, out)

    print("\nVote shares")
    xtd_shares = find_xtd_shares(xtd_votes)
    print_table(xtd_shares, header, const_names, out, "{:.1%}")

    print("\nConstituency seats")
    xtd_const_seats = add_totals(election.m_const_seats_alloc)
    print_table(xtd_const_seats, header, const_names, out)

    print("\nAdjustment seat apportionment")
    print("Threshold: {:.1%}".format(rules["adjustment_threshold"]*0.01))
    v_votes = election.v_votes
    v_votes.append(sum(election.v_votes))
    v_elim_votes = election.v_votes_eliminated
    v_elim_votes.append(sum(election.v_votes_eliminated))
    v_elim_shares = ["{:.1%}".format(v/v_elim_votes[-1])
                        for v in v_elim_votes]
    v_const_seats = election.v_const_seats_alloc
    v_const_seats.append(sum(election.v_const_seats_alloc))
    data = [v_votes, v_elim_votes, v_elim_shares, v_const_seats]
    labels = ["Total votes", "Votes above threshold",
              "Vote shares above threshold", "Constituency seats"]
    print_table(data, header[1:], labels, out)

    method = ADJUSTMENT_METHODS[rules["adjustment_method"]]
    try:
        h, data = method.print_seats(rules, election.adj_seats_info)
        print("")
        print(tabulate(data, h, out))
        print("")
    except AttributeError:
        pass

    xtd_total_seats = add_totals(election.results)
    print("\nAdjustment seats")
    xtd_adj_seats = matrix_subtraction(xtd_total_seats, xtd_const_seats)
    print_table(xtd_adj_seats, header, const_names, out)

    print("\nTotal seats")
    print_table(xtd_total_seats, header, const_names, out)

    print("\nSeat shares")
    xtd_shares = find_xtd_shares(xtd_total_seats)
    print_table(xtd_shares, header, const_names, out, "{:.1%}")

def pretty_print_election(election):
    """Print results of a single election."""
    rules = election.rules
    header = ["Constituency"]
    header.extend(rules["parties"])
    header.append("Total")
    if "constituencies" in rules:
        const_names = [c["name"] for c in rules["constituencies"]]
    else:
        const_names = rules["constituency_names"]
    const_names.append("Total")
    xtd_results = add_totals(election.results)
    print_table(xtd_results, header, const_names, rules["output"])

def entropy(votes, allocations, divisor_gen):
    """
    Calculate entropy of the election, taking into account votes and
     allocations.
     $\\sum_i \\sum_j \\sum_k \\log{v_{ij}/d_k}$, more or less.
    """
    assert(type(votes) == list)
    assert(all(type(v) == list for v in votes))
    assert(type(allocations) == list)
    assert(all(type(a) == list for a in allocations))

    e = 0
    for c in range(len(votes)):
        for p in range(len(votes[c])):
            gen = divisor_gen()
            for k in range(allocations[c][p]):
                dk = next(gen)
                e += log(votes[c][p]/dk)
    return e

def election_to_xlsx(election, filename):
    """Write detailed information about a single election to an xlsx file."""
    if "constituencies" in election.rules:
        const_names = [c["name"] for c in election.rules["constituencies"]]
    else:
        const_names = election.rules["constituency_names"]
    const_names.append("Total")
    parties = election.rules["parties"] + ["Total"]
    xtd_votes = add_totals(election.m_votes)
    xtd_shares = find_xtd_shares(xtd_votes)
    xtd_const_seats = add_totals(election.m_const_seats_alloc)
    xtd_total_seats = add_totals(election.results)
    xtd_adj_seats = matrix_subtraction(xtd_total_seats, xtd_const_seats)
    xtd_seat_shares = find_xtd_shares(xtd_total_seats)
    threshold = 0.01*election.rules["adjustment_threshold"]
    xtd_final_votes = add_totals([election.v_votes_eliminated])[0]
    xtd_final_shares = find_xtd_shares([xtd_final_votes])[0]

    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    cell_format = workbook.add_format()
    cell_format.set_align('right')
    share_format = workbook.add_format()
    share_format.set_num_format('0.0%')
    h_format = workbook.add_format()
    h_format.set_align('center')
    h_format.set_bold()
    h_format.set_font_size(14)

    worksheet.set_column('B:B', 20)

    def write_matrix(worksheet, startrow, startcol, matrix, cformat):
        for c in range(len(matrix)):
            for p in range(len(matrix[c])):
                if matrix[c][p] != 0:
                    try:
                        worksheet.write(startrow+c, startcol+p, matrix[c][p],
                                        cformat[c])
                    except TypeError:
                        worksheet.write(startrow+c, startcol+p, matrix[c][p],
                                        cformat)

    def draw_block(worksheet, row, col,
        heading, xheaders, yheaders,
        matrix,
        topleft="Constituency",
        cformat=cell_format
    ):
        if heading.endswith("shares"):
            cformat = share_format
        worksheet.merge_range(
            row, col+1,
            row, col+len(xheaders),
            heading, h_format
        )
        worksheet.write(row+1, col, topleft, cell_format)
        worksheet.write_row(row+1, col+1, xheaders, cell_format)
        worksheet.write_column(row+2, col, yheaders, cell_format)
        write_matrix(worksheet, row+2, col+1, matrix, cformat)

    startcol = 1
    startrow = 4
    tables_before = [
        {"heading": "Votes",              "matrix": xtd_votes      },
        {"heading": "Vote shares",        "matrix": xtd_shares     },
        {"heading": "Constituency seats", "matrix": xtd_const_seats},
    ]
    for table in tables_before:
        draw_block(worksheet, row=startrow, col=startcol,
            heading=table["heading"], xheaders=parties, yheaders=const_names,
            matrix=table["matrix"]
        )
        startrow += 3 + len(const_names)

    row_headers = ['Total votes', 'Vote shares', 'Threshold',
                   'Votes above threshold',
                   'Vote shares above threshold', 'Constituency seats']
    matrix = [xtd_votes[-1],   xtd_shares[-1],   [threshold],
              xtd_final_votes, xtd_final_shares, xtd_const_seats[-1]]
    formats = [cell_format, share_format, share_format,
               cell_format, share_format, cell_format]
    draw_block(worksheet, row=startrow, col=startcol,
        heading="Adjustment seat apportionment", topleft="Party",
        xheaders=parties, yheaders=row_headers,
        matrix=matrix, cformat=formats
    )
    startrow += 3 + len(row_headers)

    method = ADJUSTMENT_METHODS[election.rules["adjustment_method"]]
    try:
        h, data = method.print_seats(election.rules, election.adj_seats_info)
        worksheet.merge_range(
            startrow, startcol+1,
            startrow, startcol+len(parties),
            "Step-by-step demonstration", h_format
        )
        worksheet.write_row(startrow+1, 1, h, cell_format)
        for i in range(len(data)):
            worksheet.write_row(startrow+2+i, 1, data[i], cell_format)
        startrow += 3 + len(data)
    except AttributeError:
        pass

    tables_after = [
        {"heading": "Adjustment seats", "matrix": xtd_adj_seats  },
        {"heading": "Total seats",      "matrix": xtd_total_seats},
        {"heading": "Seat shares",      "matrix": xtd_seat_shares},
    ]
    for table in tables_after:
        draw_block(worksheet, row=startrow, col=startcol,
            heading=table["heading"], xheaders=parties, yheaders=const_names,
            matrix=table["matrix"]
        )
        startrow += 3 + len(const_names)

    worksheet.write(startrow, startcol, 'Entropy:', h_format)
    worksheet.write(startrow, startcol+1, election.entropy(), cell_format)

    workbook.close()

def sim_election_rules(rs, test_method):
    """Get preset election rules for simulation from file."""
    config = configparser.ConfigParser()
    config.read("../data/presets/methods.ini")

    if test_method in config:
        rs.update(config[test_method])
    else:
        raise ValueError("%s is not a known apportionment method"
                            % test_method)
    rs["adjustment_threshold"] = float(rs["adjustment_threshold"])

    return rs

def print_simulation(simulation):
    """Print detailed information about a simulation."""
    for r in range(len(simulation.e_rules)):
        rules = simulation.e_rules[r]
        out = rules["output"]
        print("==========================================")
        print("Adjustment method:", rules["adjustment_method"])
        print("==========================================\n")
        h = ["Constituency"]
        h.extend(rules["parties"]+["Total"])
        if "constituencies" in rules:
            const_names = [c["name"] for c in rules["constituencies"]]
        else:
            const_names = rules["constituency_names"]
        const_names.append("Total")

        print("Reference")

        print("\nVotes")
        print_table(simulation.xtd_votes, h, const_names, out)

        print("\nVote shares")
        print_table(simulation.xtd_vote_shares, h, const_names, out, "{:.1%}")

        print("\nConstituency seats")
        print_table(simulation.base_allocations[r]["xtd_const_seats"], h, const_names, out)

        print("\nAdjustment seats")
        print_table(simulation.base_allocations[r]["xtd_adj_seats"], h, const_names, out)

        print("\nTotal seats")
        print_table(simulation.base_allocations[r]["xtd_total_seats"], h, const_names, out)

        print("\nSeat shares")
        print_table(simulation.base_allocations[r]["xtd_seat_shares"], h, const_names, out, "{:.1%}")

        print("\nAverages from simulation")

        print("\nVotes")
        print_table(simulation.list_data[-1]["sim_votes"]["avg"], h, const_names, out)

        print("\nVote shares")
        print_table(simulation.list_data[-1]["sim_shares"]["avg"], h, const_names, out, "{:.1%}")

        print("\nConstituency seats")
        print_table(simulation.list_data[r]["const_seats"]["avg"], h, const_names, out)

        print("\nAdjustment seats")
        print_table(simulation.list_data[r]["adj_seats"]["avg"], h, const_names, out)

        print("\nTotal seats")
        print_table(simulation.list_data[r]["total_seats"]["avg"], h, const_names, out)

        print("\nSeat shares")
        print_table(simulation.list_data[r]["seat_shares"]["avg"], h, const_names, out, "{:.1%}")

        print("\nStandard deviations from simulation")

        print("\nVotes")
        print_table(simulation.list_data[-1]["sim_votes"]["std"], h, const_names, out, "{:.3f}")

        print("\nVote shares")
        print_table(simulation.list_data[-1]["sim_shares"]["std"], h, const_names, out, "{:.1%}")

        print("\nConstituency seats")
        print_table(simulation.list_data[r]["const_seats"]["std"], h, const_names, out, "{:.3f}")

        print("\nAdjustment seats")
        print_table(simulation.list_data[r]["adj_seats"]["std"], h, const_names, out, "{:.3f}")

        print("\nTotal seats")
        print_table(simulation.list_data[r]["total_seats"]["std"], h, const_names, out, "{:.3f}")

        print("\nSeat shares")
        print_table(simulation.list_data[r]["seat_shares"]["std"], h, const_names, out, "{:.1%}")

        #print("\nVotes added to change results")
        #print_table(simulation.votes_to_change, h[:-1], const_names[:-1], out)

def simulation_to_xlsx(simulation, filename):
    """Write detailed information about a simulation to an xlsx file."""
    workbook = xlsxwriter.Workbook(filename)

    r_format = workbook.add_format()
    r_format.set_rotation(90)
    r_format.set_align('center')
    r_format.set_align('vcenter')
    r_format.set_text_wrap()
    r_format.set_bold()
    r_format.set_font_size(12)

    h_format = workbook.add_format()
    h_format.set_align('center')
    h_format.set_bold()
    h_format.set_font_size(14)

    cell_format = workbook.add_format()
    cell_format.set_align('right')

    base_format = workbook.add_format()
    base_format.set_num_format('#,##0')

    sim_format = workbook.add_format()
    sim_format.set_num_format('#,##0.000')

    share_format = workbook.add_format()
    share_format.set_num_format('0.0%')


    def write_matrix(worksheet, startrow, startcol, matrix, cformat):
        for c in range(len(matrix)):
            for p in range(len(matrix[c])):
                if matrix[c][p] != 0:
                    worksheet.write(startrow+c, startcol+p, matrix[c][p],
                                    cformat)

    def draw_block(worksheet, row, col,
        heading, xheaders, yheaders,
        matrix,
        cformat=cell_format
    ):
        if heading.endswith("shares"):
            cformat = share_format
        if heading == "Votes":
            cformat = base_format
        worksheet.merge_range(
            row, col, row, col+len(xheaders), heading, h_format)
        worksheet.write_row(row+1, col+1, xheaders, cell_format)
        worksheet.write_column(row+2, col, yheaders, cell_format)
        write_matrix(worksheet, row+2, col+1, matrix, cformat)

    categories = [
        {"abbr": "base", "cell_format": base_format,
         "heading": "Reference data"                     },
        {"abbr": "avg",  "cell_format": sim_format,
         "heading": "Averages from simulation"           },
        {"abbr": "std",  "cell_format": sim_format,
         "heading": "Standard deviations from simulation"},
    ]
    tables = [
        {"abbr": "v",  "heading": "Votes"             },
        {"abbr": "vs", "heading": "Vote shares"       },
        {"abbr": "cs", "heading": "Constituency seats"},
        {"abbr": "as", "heading": "Adjustment seats"  },
        {"abbr": "ts", "heading": "Total seats"       },
        {"abbr": "ss", "heading": "Seat shares"       },
    ]

    for r in range(len(simulation.e_rules)):
        method_name = simulation.e_rules[r]["adjustment_method"]
        worksheet   = workbook.add_worksheet(method_name)
        const_names = simulation.e_rules[r]["constituency_names"] + ["Total"]
        parties     = simulation.e_rules[r]["parties"           ] + ["Total"]

        data_matrix = {
            "base": {
                "v" : simulation.xtd_votes,
                "vs": simulation.xtd_vote_shares,
                "cs": simulation.base_allocations[r]["xtd_const_seats"],
                "as": simulation.base_allocations[r]["xtd_adj_seats"],
                "ts": simulation.base_allocations[r]["xtd_total_seats"],
                "ss": simulation.base_allocations[r]["xtd_seat_shares"],
            },
            "avg": {
                "v" : simulation.list_data[-1]["sim_votes"  ]["avg"],
                "vs": simulation.list_data[-1]["sim_shares" ]["avg"],
                "cs": simulation.list_data[ r]["const_seats"]["avg"],
                "as": simulation.list_data[ r]["adj_seats"  ]["avg"],
                "ts": simulation.list_data[ r]["total_seats"]["avg"],
                "ss": simulation.list_data[ r]["seat_shares"]["avg"],
            },
            "std": {
                "v" : simulation.list_data[-1]["sim_votes"  ]["std"],
                "vs": simulation.list_data[-1]["sim_shares" ]["std"],
                "cs": simulation.list_data[ r]["const_seats"]["std"],
                "as": simulation.list_data[ r]["adj_seats"  ]["std"],
                "ts": simulation.list_data[ r]["total_seats"]["std"],
                "ss": simulation.list_data[ r]["seat_shares"]["std"],
            },
        }
        toprow = 3
        for category in categories:
            worksheet.merge_range(toprow, 0, toprow+1+len(const_names), 0,
                category["heading"], r_format)
            col = 2
            for table in tables:
                draw_block(worksheet, row=toprow, col=col,
                    heading=table["heading"],
                    xheaders=parties,
                    yheaders=const_names,
                    matrix=data_matrix[category["abbr"]][table["abbr"]],
                    cformat=category["cell_format"]
                )
                col += len(parties)+2
            toprow += len(const_names)+3

        results = simulation.get_results_dict()
        DEVIATION_MEASURES = results["deviation_measures"]
        STANDARDIZED_MEASURES = results["standardized_measures"]
        MEASURES = results["measures"]
        mkeys = DEVIATION_MEASURES + STANDARDIZED_MEASURES
        measure_names = [MEASURES[key] for key in mkeys]
        aggregates = ["avg", "std"]
        aggregate_names = [results["aggregates"][aggr] for aggr in aggregates]
        measure_table = [
            [simulation.data[r][measure][aggr] for aggr in aggregates]
            for measure in mkeys
        ]
        draw_block(worksheet, row=toprow, col=9,
            heading="Summary measures",
            xheaders=aggregate_names,
            yheaders=measure_names,
            matrix=measure_table,
            cformat=sim_format
        )

    workbook.close()

ADJUSTMENT_METHODS = {
    "alternating-scaling" : alternating_scaling,
    "relative-superiority": relative_superiority,
    "nearest-neighbor"    : nearest_neighbor,
    "monge"               : monge,
    "icelandic-law"       : icelandic_law,
    "norwegian-law"       : norwegian_law,
    "norwegian-icelandic" : norwegian_icelandic,
    "opt-entropy"         : opt_entropy,
    "switching"           : switching,
    "var-alt-scal"        : var_alt_scal,
    "pure-vote-ratios"    : pure_vote_ratios,
}
